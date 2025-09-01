#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PyQt clone of the MATLAB 'CellAreaEstimator' app.

This script provides a GUI built using **PyQt6** (the Qt for Python bindings
available via the ``PyQt6`` package).  It mirrors the functionality of the
original MATLAB application: loading difference and binary masks, selecting
input/output folders, configuring registration and segmentation parameters,
processing all images in the input directory, previewing intermediate
results, and saving both the processed images and cell area measurements.

**Why PyQt6?**  Modern PyQt releases require Python 3.8 or newer.  In fact,
the PyPI metadata for PyQt5 notes that it "requires Python >=3.8"【566249812113657†L60-L65】,
and the PyQt6 package requires Python >=3.9【489455742411333†L59-L63】.  Earlier versions like
PyQt4 are long‑unmaintained and only support very old Python releases【5997990753853†L97-L100】.
Therefore, there is no supported PyQt version for Python 3.1, so this
implementation targets Python 3.9+ with PyQt6.

To run this script you will need the following third‑party packages:

  * PyQt6 (``pip install PyQt6``)
  * OpenCV (``pip install opencv-python``)
  * NumPy (``pip install numpy``)
  * pandas (``pip install pandas``)
  * openpyxl (``pip install openpyxl``)
  * scikit‑image (optional, for Chan–Vese segmentation; ``pip install scikit-image``)

The application intentionally avoids any platform‑specific APIs so it should
work on Windows, macOS and Linux provided the above dependencies are
installed.  As with the PySide6 variant, this environment does not have
a graphical toolkit available, so you will need to run the script on your
own machine to use the GUI.

"""

from __future__ import annotations

import os
import sys
import time
from dataclasses import dataclass
from typing import List, Optional

import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

try:
    from skimage.segmentation import morphological_chan_vese
    SKIMAGE_AVAILABLE = True
except Exception:
    SKIMAGE_AVAILABLE = False

# PyQt6 imports
from PyQt6.QtCore import (Qt, QObject, QThread, pyqtSignal, pyqtSlot)
from PyQt6.QtGui import QImage, QPixmap
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QLineEdit,
    QFileDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QSpinBox, QDoubleSpinBox, QCheckBox, QDialog, QMessageBox, QProgressBar,
    QAction
)


# -----------------------------------------------------------------------------
# Utility functions
#
def imread_gray(path: str) -> np.ndarray:
    """Read an image from *path* and ensure it is grayscale."""
    img = cv2.imread(path, cv2.IMREAD_UNCHANGED)
    if img is None:
        raise IOError(f"Failed to read image: {path}")
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return img


def to_uint8(img: np.ndarray) -> np.ndarray:
    """Normalize an array to uint8 (0–255)."""
    if img.dtype == np.uint8:
        return img
    img = np.asarray(img, dtype=np.float32)
    mn, mx = img.min(), img.max()
    if mx - mn < 1e-8:
        return np.zeros_like(img, dtype=np.uint8)
    out = (img - mn) / (mx - mn) * 255.0
    return np.clip(out, 0, 255).astype(np.uint8)


def qimage_from_gray(img: np.ndarray) -> QImage:
    """Convert a grayscale array to a QImage."""
    g = to_uint8(img)
    h, w = g.shape
    return QImage(g.data, w, h, w, QImage.Format.Format_Grayscale8)


def ksize_from_sigma(sigma: float) -> int:
    """Given a Gaussian sigma, compute an odd kernel size for OpenCV."""
    if sigma <= 0:
        return 0
    k = int(6 * sigma + 1)
    return k if k % 2 == 1 else k + 1


def clahe_equalize(img: np.ndarray) -> np.ndarray:
    """Apply contrast limited adaptive histogram equalization."""
    img8 = to_uint8(img)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    return clahe.apply(img8)


def complement(img: np.ndarray) -> np.ndarray:
    """Return the complement (inverted) of a grayscale image."""
    return 255 - to_uint8(img)


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def list_jpgs(folder: str) -> List[str]:
    return [f for f in os.listdir(folder) if f.lower().endswith('.jpg')]


def num2xlcol(col_num: int) -> str:
    """Convert a 1‑based column index to an Excel column name (A, B, ...)."""
    n = col_num
    col_chars = []
    while n > 0:
        n, r = divmod(n - 1, 26)
        col_chars.append(chr(r + 65))
    return ''.join(reversed(col_chars))


def write_sorted_areas_xlsx(xlsx_path: str, column_index: int, areas: List[int]) -> None:
    """Write a list of integer areas into *xlsx_path* at column *column_index* (1‑based)."""
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    row = 1
    for area in areas:
        ws.cell(row=row, column=column_index, value=int(area))
        row += 1
    wb.save(xlsx_path)


def connected_component_areas(binary: np.ndarray) -> List[int]:
    """Compute the areas of connected components in a binary mask (excluding background)."""
    b = (binary > 0).astype(np.uint8)
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(b, connectivity=8)
    return [int(stats[i, cv2.CC_STAT_AREA]) for i in range(1, num_labels)]


# -----------------------------------------------------------------------------
# Registration and segmentation
#
@dataclass
class RegSegParams:
    """Container for registration and segmentation parameters."""
    # Registration (ECC approximation)
    numSpaSam: int = 5000       # not directly used by ECC but preserved for UI
    numHisBin: int = 64         # not used by ECC
    allPix: bool = True         # not used by ECC
    groFac: float = 1.5         # not used by ECC
    epsilon: float = 1e-6       # ECC termination epsilon
    iniRad: float = 1.0         # not used by ECC
    maxIter: int = 200          # ECC max iterations
    gausBlurDif: float = 2.0    # Gaussian sigma for difference mask
    gausBlurIn: float = 2.0     # Gaussian sigma for input image
    bwThresh: int = 50          # threshold (>=0) or Otsu (<0)
    recMaSize: int = 3          # rectangular opening kernel size
    segIter: int = 50           # Chan–Vese iterations if available


def register_ecc(moving: np.ndarray, fixed: np.ndarray, params: RegSegParams) -> np.ndarray:
    """Register *moving* to *fixed* using OpenCV ECC with an affine model."""
    # Pre‑blur both images (if sigma > 0)
    if params.gausBlurDif > 0:
        kf = ksize_from_sigma(params.gausBlurDif)
        fixed_blur = cv2.GaussianBlur(fixed, (kf, kf), params.gausBlurDif) if kf > 0 else fixed
    else:
        fixed_blur = fixed
    if params.gausBlurIn > 0:
        km = ksize_from_sigma(params.gausBlurIn)
        moving_blur = cv2.GaussianBlur(moving, (km, km), params.gausBlurIn) if km > 0 else moving
    else:
        moving_blur = moving
    # Normalize to float32 [0,1]
    fb = to_uint8(fixed_blur).astype(np.float32) / 255.0
    mb = to_uint8(moving_blur).astype(np.float32) / 255.0
    # Initialize warp matrix for affine transform
    warp_matrix = np.eye(2, 3, dtype=np.float32)
    criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, params.maxIter, params.epsilon)
    try:
        cv2.findTransformECC(fb, mb, warp_matrix, cv2.MOTION_AFFINE, criteria)
    except cv2.error:
        # If ECC fails, keep identity warp
        pass
    h, w = fixed.shape
    registered = cv2.warpAffine(moving, warp_matrix, (w, h), flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_REPLICATE)
    return registered


def segment_image(img: np.ndarray, params: RegSegParams) -> np.ndarray:
    """Segment an image using thresholding, morphology and optional Chan–Vese."""
    g = to_uint8(img)
    # Thresholding: fixed threshold or Otsu
    if params.bwThresh >= 0:
        _, BW = cv2.threshold(g, int(params.bwThresh), 255, cv2.THRESH_BINARY)
    else:
        _, BW = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # Fill holes: invert, flood fill from border, subtract to get holes
    inv = 255 - BW
    h, w = inv.shape
    mask = np.zeros((h + 2, w + 2), dtype=np.uint8)
    flood = inv.copy()
    cv2.floodFill(flood, mask, (0, 0), 0)
    holes = inv - flood
    BW_filled = BW.copy()
    BW_filled[holes > 0] = 255
    # Morphological opening with rectangular kernel
    k = max(1, int(params.recMaSize))
    rect = cv2.getStructuringElement(cv2.MORPH_RECT, (k, k))
    BW_open = cv2.morphologyEx(BW_filled, cv2.MORPH_OPEN, rect, iterations=1)
    # Optional Chan–Vese refinement
    if SKIMAGE_AVAILABLE and params.segIter > 0:
        img01 = g.astype(np.float32) / 255.0
        init_ls = BW_open.astype(bool)
        cv_res = morphological_chan_vese(img01, iterations=params.segIter, init_level_set=init_ls, smoothing=1)
        BW_refined = (cv_res.astype(np.uint8) * 255)
    else:
        BW_refined = BW_open
    return BW_refined


# -----------------------------------------------------------------------------
# Worker for background processing
#
class ProcessorWorker(QObject):
    """Worker object that processes images in a separate thread."""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    imagePreviews = pyqtSignal(np.ndarray, np.ndarray, np.ndarray)
    diffPreviews = pyqtSignal(np.ndarray, np.ndarray)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, in_dir: str, out_dir: str, dm_path: str, bm_path: str, params: RegSegParams):
        super().__init__()
        self.in_dir = in_dir
        self.out_dir = out_dir
        self.dm_path = dm_path
        self.bm_path = bm_path
        self.params = params
        self._stop = False
        self._pause = False

    def stop(self) -> None:
        self._stop = True

    def toggle_pause(self, value: bool) -> None:
        self._pause = value

    def _check_pause_stop(self) -> None:
        while self._pause and not self._stop:
            time.sleep(0.2)
        if self._stop:
            raise RuntimeError("Processing stopped by user.")

    @pyqtSlot()
    def run(self) -> None:
        """Entry point for the worker thread."""
        try:
            # Load masks once
            dm = imread_gray(self.dm_path)
            bm = imread_gray(self.bm_path)
            # Emit initial preview (placeholder current image is difference mask)
            self.imagePreviews.emit(dm, bm, dm)
            # Ensure output subdirectories exist
            top_dir = os.path.join(self.out_dir, "top")
            topbw_dir = os.path.join(self.out_dir, "topBW")
            bot_dir = os.path.join(self.out_dir, "bottom")
            botbw_dir = os.path.join(self.out_dir, "bottomBW")
            for d in (top_dir, topbw_dir, bot_dir, botbw_dir):
                ensure_dir(d)
            # Precompute mask differences for top/bottom
            binDif_top = cv2.subtract(dm, bm)
            binDif_bot = cv2.subtract(dm, complement(bm))
            files = list_jpgs(self.in_dir)
            if not files:
                self.error.emit("No .jpg files found in the input directory.")
                self.finished.emit()
                return
            total = len(files)
            top_xlsx = os.path.join(self.out_dir, "top.xlsx")
            bot_xlsx = os.path.join(self.out_dir, "bottom.xlsx")
            for idx, fname in enumerate(files, start=1):
                self._check_pause_stop()
                self.status.emit(f"Processing {idx}/{total}: {fname}")
                path = os.path.join(self.in_dir, fname)
                cur = imread_gray(path)
                # Update previews
                self.imagePreviews.emit(dm, bm, cur)
                # Register current image to difference mask
                reg = register_ecc(cur, dm, self.params)
                # --- TOP ---
                binReg_top = cv2.subtract(reg, bm)
                topDiff = cv2.subtract(clahe_equalize(binDif_top), clahe_equalize(binReg_top))
                # Emit top diff preview first (bottom diff will be emitted second)
                self.diffPreviews.emit(topDiff, topDiff)
                cv2.imwrite(os.path.join(top_dir, fname), to_uint8(topDiff))
                topBW = segment_image(topDiff, self.params)
                cv2.imwrite(os.path.join(topbw_dir, fname), to_uint8(topBW))
                areas_top = connected_component_areas(topBW)
                areas_top.sort(reverse=True)
                write_sorted_areas_xlsx(top_xlsx, idx, areas_top)
                # --- BOTTOM ---
                binReg_bot = cv2.subtract(reg, complement(bm))
                botDiff = cv2.subtract(clahe_equalize(binDif_bot), clahe_equalize(binReg_bot))
                # Update diff previews with both top and bottom
                self.diffPreviews.emit(topDiff, botDiff)
                cv2.imwrite(os.path.join(bot_dir, fname), to_uint8(botDiff))
                botBW = segment_image(botDiff, self.params)
                cv2.imwrite(os.path.join(botbw_dir, fname), to_uint8(botBW))
                areas_bot = connected_component_areas(botBW)
                areas_bot.sort(reverse=True)
                write_sorted_areas_xlsx(bot_xlsx, idx, areas_bot)
                # Update progress
                progress_pct = int(idx / total * 100)
                self.progress.emit(progress_pct)
            self.status.emit("Done.")
            self.progress.emit(100)
            self.finished.emit()
        except Exception as exc:
            self.error.emit(str(exc))
            self.finished.emit()


# -----------------------------------------------------------------------------
# Parameter dialog
#
class ParamDialog(QDialog):
    """Dialog window for editing registration/segmentation parameters."""
    def __init__(self, params: RegSegParams, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Registration/Segmentation Parameters")
        self.params = params
        form = QFormLayout(self)
        # Registration controls
        self.sp_numSpaSam = QSpinBox(); self.sp_numSpaSam.setRange(1, 1_000_000); self.sp_numSpaSam.setValue(params.numSpaSam)
        self.sp_numHisBin = QSpinBox(); self.sp_numHisBin.setRange(1, 4096); self.sp_numHisBin.setValue(params.numHisBin)
        self.cb_allPix = QCheckBox(); self.cb_allPix.setChecked(params.allPix)
        self.dsp_groFac = QDoubleSpinBox(); self.dsp_groFac.setRange(0.0, 10.0); self.dsp_groFac.setSingleStep(0.1); self.dsp_groFac.setValue(params.groFac)
        self.dsp_epsilon = QDoubleSpinBox(); self.dsp_epsilon.setDecimals(10); self.dsp_epsilon.setRange(1e-10, 1e-1); self.dsp_epsilon.setValue(params.epsilon)
        self.dsp_iniRad = QDoubleSpinBox(); self.dsp_iniRad.setRange(0.0, 50.0); self.dsp_iniRad.setSingleStep(0.1); self.dsp_iniRad.setValue(params.iniRad)
        self.sp_maxIter = QSpinBox(); self.sp_maxIter.setRange(1, 10000); self.sp_maxIter.setValue(params.maxIter)
        form.addRow("[Reg] Num Spatial Samples", self.sp_numSpaSam)
        form.addRow("[Reg] Num Hist Bins", self.sp_numHisBin)
        form.addRow("[Reg] Use All Pixels", self.cb_allPix)
        form.addRow("[Reg] Growth Factor", self.dsp_groFac)
        form.addRow("[Reg] Epsilon", self.dsp_epsilon)
        form.addRow("[Reg] Initial Radius", self.dsp_iniRad)
        form.addRow("[Reg] Max Iterations", self.sp_maxIter)
        # Blur controls
        self.dsp_gbd = QDoubleSpinBox(); self.dsp_gbd.setRange(0.0, 50.0); self.dsp_gbd.setSingleStep(0.1); self.dsp_gbd.setValue(params.gausBlurDif)
        self.dsp_gbi = QDoubleSpinBox(); self.dsp_gbi.setRange(0.0, 50.0); self.dsp_gbi.setSingleStep(0.1); self.dsp_gbi.setValue(params.gausBlurIn)
        form.addRow("[Blur] Gaussian sigma (DM)", self.dsp_gbd)
        form.addRow("[Blur] Gaussian sigma (Input)", self.dsp_gbi)
        # Segmentation controls
        self.sp_thresh = QSpinBox(); self.sp_thresh.setRange(-1, 255); self.sp_thresh.setValue(params.bwThresh)
        self.sp_rect = QSpinBox(); self.sp_rect.setRange(1, 99); self.sp_rect.setSingleStep(1); self.sp_rect.setValue(params.recMaSize)
        self.sp_segIter = QSpinBox(); self.sp_segIter.setRange(0, 10000); self.sp_segIter.setValue(params.segIter)
        form.addRow("[Seg] Threshold (−1=Otsu)", self.sp_thresh)
        form.addRow("[Seg] Rect Mask Size", self.sp_rect)
        form.addRow("[Seg] Chan–Vese Iterations", self.sp_segIter)
        # Buttons
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("OK"); ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Cancel"); cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        form.addRow(btn_layout)

    def apply(self) -> None:
        """Update the bound RegSegParams with the current UI values."""
        self.params.numSpaSam = int(self.sp_numSpaSam.value())
        self.params.numHisBin = int(self.sp_numHisBin.value())
        self.params.allPix = bool(self.cb_allPix.isChecked())
        self.params.groFac = float(self.dsp_groFac.value())
        self.params.epsilon = float(self.dsp_epsilon.value())
        self.params.iniRad = float(self.dsp_iniRad.value())
        self.params.maxIter = int(self.sp_maxIter.value())
        self.params.gausBlurDif = float(self.dsp_gbd.value())
        self.params.gausBlurIn = float(self.dsp_gbi.value())
        self.params.bwThresh = int(self.sp_thresh.value())
        self.params.recMaSize = int(self.sp_rect.value())
        self.params.segIter = int(self.sp_segIter.value())


# -----------------------------------------------------------------------------
# Main window
#
class MainWindow(QMainWindow):
    """The main window of the cell area estimator."""
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Cell Area Estimator (PyQt6)")
        self.params = RegSegParams()
        # Menu
        menubar = self.menuBar()
        tools_menu = menubar.addMenu("Tools")
        params_action = QAction("Registration/Segmentation Setup", self)
        tools_menu.addAction(params_action)
        params_action.triggered.connect(self.open_param_dialog)
        # File selection controls
        self.dm_label = QLabel("Difference Mask:")
        self.dm_path = QLineEdit(); self.dm_path.setReadOnly(True)
        self.btn_dm = QPushButton("Load DM"); self.btn_dm.clicked.connect(self.load_dm)
        self.bm_label = QLabel("Binary Mask:")
        self.bm_path = QLineEdit(); self.bm_path.setReadOnly(True)
        self.btn_bm = QPushButton("Load BM"); self.btn_bm.clicked.connect(self.load_bm)
        self.in_label = QLabel("Input Image Directory:")
        self.in_path = QLineEdit(); self.in_path.setReadOnly(True)
        self.btn_in = QPushButton("Select Dir"); self.btn_in.clicked.connect(self.select_input_dir)
        self.out_label = QLabel("Save Difference Images To Directory:")
        self.out_path = QLineEdit(); self.out_path.setReadOnly(True)
        self.btn_out = QPushButton("Select Dir"); self.btn_out.clicked.connect(self.select_output_dir)
        # Buttons for processing and control
        self.btn_process = QPushButton("Process (Single Type)")
        self.btn_process.clicked.connect(self.start_processing)
        self.btn_pause = QPushButton("Pause")
        self.btn_pause.setCheckable(True)
        self.btn_pause.toggled.connect(self.toggle_pause)
        self.btn_stop = QPushButton("Stop")
        self.btn_stop.clicked.connect(self.stop_processing)
        self.btn_stop.setStyleSheet("color: red; font-weight: bold;")
        # Progress
        self.progress_label = QLabel("Current Progress")
        self.progress_bar = QProgressBar(); self.progress_bar.setRange(0, 100)
        # Image previews
        self.img_dm = QLabel(); self.img_dm.setMinimumSize(240, 240); self.img_dm.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_bm = QLabel(); self.img_bm.setMinimumSize(240, 240); self.img_bm.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_input = QLabel(); self.img_input.setMinimumSize(240, 240); self.img_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_top = QLabel(); self.img_top.setMinimumSize(240, 240); self.img_top.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_bottom = QLabel(); self.img_bottom.setMinimumSize(240, 240); self.img_bottom.setAlignment(Qt.AlignmentFlag.AlignCenter)
        # Layouts
        meta_layout = QGridLayout()
        meta_layout.addWidget(self.dm_label, 0, 0); meta_layout.addWidget(self.dm_path, 0, 1); meta_layout.addWidget(self.btn_dm, 0, 2)
        meta_layout.addWidget(self.bm_label, 1, 0); meta_layout.addWidget(self.bm_path, 1, 1); meta_layout.addWidget(self.btn_bm, 1, 2)
        meta_layout.addWidget(self.in_label, 2, 0); meta_layout.addWidget(self.in_path, 2, 1); meta_layout.addWidget(self.btn_in, 2, 2)
        meta_layout.addWidget(self.out_label, 3, 0); meta_layout.addWidget(self.out_path, 3, 1); meta_layout.addWidget(self.btn_out, 3, 2)
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.btn_process)
        btn_layout.addWidget(self.progress_label)
        btn_layout.addWidget(self.progress_bar)
        btn_layout.addWidget(self.btn_pause)
        btn_layout.addWidget(self.btn_stop)
        grid_layout = QGridLayout()
        grid_layout.addWidget(self.img_dm, 0, 0)
        grid_layout.addWidget(self.img_bm, 0, 1)
        grid_layout.addWidget(self.img_input, 1, 0)
        grid_layout.addWidget(self.img_top, 1, 1)
        grid_layout.addWidget(self.img_bottom, 1, 2)
        central = QWidget()
        vbox = QVBoxLayout(central)
        vbox.addLayout(meta_layout)
        vbox.addLayout(btn_layout)
        vbox.addLayout(grid_layout)
        self.setCentralWidget(central)
        # Worker fields
        self.worker_thread: Optional[QThread] = None
        self.worker: Optional[ProcessorWorker] = None

    # ----- Slots -----
    def open_param_dialog(self) -> None:
        dlg = ParamDialog(self.params, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            dlg.apply()

    def load_dm(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Select Difference Mask", "", "Images (*.jpg *.jpeg *.png *.tif *.tiff)")
        if path:
            self.dm_path.setText(path)
            try:
                im = imread_gray(path)
                self.set_label_image(self.img_dm, im)
            except Exception as exc:
                QMessageBox.critical(self, "Error", str(exc))

    def load_bm(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Select Binary Mask", "", "Images (*.jpg *.jpeg *.png *.tif *.tiff)")
        if path:
            self.bm_path.setText(path)
            try:
                im = imread_gray(path)
                self.set_label_image(self.img_bm, im)
            except Exception as exc:
                QMessageBox.critical(self, "Error", str(exc))

    def select_input_dir(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "Select Input Directory")
        if directory:
            self.in_path.setText(directory)

    def select_output_dir(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "Select Output Directory")
        if directory:
            self.out_path.setText(directory)

    def start_processing(self) -> None:
        in_dir = self.in_path.text().strip()
        out_dir = self.out_path.text().strip()
        dm = self.dm_path.text().strip()
        bm = self.bm_path.text().strip()
        if not (os.path.isdir(in_dir) and os.path.isdir(out_dir) and os.path.isfile(dm) and os.path.isfile(bm)):
            QMessageBox.warning(self, "Missing Input", "Please provide valid DM, BM, input directory and output directory.")
            return
        self.progress_bar.setValue(0)
        # Prepare worker and thread
        self.worker_thread = QThread()
        self.worker = ProcessorWorker(in_dir, out_dir, dm, bm, self.params)
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        # Connect signals
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.status.connect(self.statusBar().showMessage)
        self.worker.imagePreviews.connect(self.update_image_previews)
        self.worker.diffPreviews.connect(self.update_diff_previews)
        self.worker.error.connect(lambda msg: QMessageBox.critical(self, "Error", msg))
        # Start thread
        self.worker_thread.start()

    def toggle_pause(self, checked: bool) -> None:
        if self.worker:
            self.worker.toggle_pause(checked)
        self.btn_pause.setText("Resume" if checked else "Pause")

    def stop_processing(self) -> None:
        if self.worker:
            self.worker.stop()

    @pyqtSlot(np.ndarray, np.ndarray, np.ndarray)
    def update_image_previews(self, dm: np.ndarray, bm: np.ndarray, cur: np.ndarray) -> None:
        self.set_label_image(self.img_dm, dm)
        self.set_label_image(self.img_bm, bm)
        self.set_label_image(self.img_input, cur)

    @pyqtSlot(np.ndarray, np.ndarray)
    def update_diff_previews(self, top: np.ndarray, bottom: np.ndarray) -> None:
        self.set_label_image(self.img_top, top)
        self.set_label_image(self.img_bottom, bottom)

    def set_label_image(self, label: QLabel, img: np.ndarray) -> None:
        qimg = qimage_from_gray(img)
        pix = QPixmap.fromImage(qimg)
        label.setPixmap(pix.scaled(label.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))


# -----------------------------------------------------------------------------
# Program entry point
#
def main() -> None:
    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(1200, 800)
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()