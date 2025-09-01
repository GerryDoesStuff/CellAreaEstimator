from __future__ import annotations

from pathlib import Path
from typing import List

import cv2
import numpy as np
import logging
from openpyxl import Workbook, load_workbook
from PyQt6.QtGui import QImage

logger = logging.getLogger(__name__)


def imread_gray(path: Path | str) -> np.ndarray:
    """Read an image from *path* and ensure it is grayscale."""
    p = Path(path)
    img = cv2.imread(str(p), cv2.IMREAD_UNCHANGED)
    if img is None:
        logger.error("Failed to read image: %s", p)
        raise IOError(f"Failed to read image: {p}")
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return img


def to_uint8(img: np.ndarray) -> np.ndarray:
    """Normalize an array to uint8 (0–255)."""
    if img.dtype == np.uint8:
        return img
    img = np.asarray(img, dtype=np.float32)
    mn, mx = img.min(), img.max()
    if mx - mn < 1e-8:
        return np.zeros_like(img, dtype=np.uint8)
    out = (img - mn) / (mx - mn) * 255.0
    return np.clip(out, 0, 255).astype(np.uint8)


def qimage_from_gray(img: np.ndarray) -> QImage:
    """Convert a grayscale array to a QImage."""
    g = to_uint8(img)
    h, w = g.shape
    return QImage(g.data, w, h, w, QImage.Format.Format_Grayscale8)


def ensure_dir(path: Path | str) -> None:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    logger.debug("Ensured directory exists: %s", p)


def list_jpgs(folder: Path | str) -> List[Path]:
    p = Path(folder)
    files = sorted(
        (f.resolve() for f in p.glob("*.{jpg,jpeg,png,tif,tiff}", case_sensitive=False))
    )
    logger.debug("Found %d image files in %s", len(files), p)
    return files


def num2xlcol(col_num: int) -> str:
    """Convert a 1‑based column index to an Excel column name (A, B, ...)."""
    n = col_num
    col_chars = []
    while n > 0:
        n, r = divmod(n - 1, 26)
        col_chars.append(chr(r + 65))
    return ''.join(reversed(col_chars))


def write_sorted_areas_xlsx(xlsx_path: Path | str, column_index: int, areas: List[int]) -> None:
    """Write a list of integer areas into *xlsx_path* at column *column_index* (1‑based)."""
    path = Path(xlsx_path)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    row = 1
    for area in areas:
        ws.cell(row=row, column=column_index, value=int(area))
        row += 1
    wb.save(path)
