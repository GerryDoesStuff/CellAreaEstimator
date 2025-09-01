from __future__ import annotations

from pathlib import Path
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Sequence

import cv2
import numpy as np
from PyQt6.QtCore import (
    QObject,
    pyqtSignal,
    pyqtSlot,
    QMutex,
    QWaitCondition,
)

from io_utils import (
    imread_gray,
    ensure_dir,
    to_uint8,
)
from processing import (
    RegSegParams,
    register_ecc,
    segment_image,
    connected_component_areas,
    clahe_equalize,
    complement,
)
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def _apply_roi(
    dm: np.ndarray,
    bm: np.ndarray,
    binDif_top: np.ndarray,
    binDif_bot: np.ndarray,
    roi: tuple[int, int, int, int] | None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Return baseline arrays cropped to *roi* if provided."""
    if roi is None:
        return dm, bm, binDif_top, binDif_bot
    x, y, w, h = roi
    return (
        dm[y:y + h, x:x + w],
        bm[y:y + h, x:x + w],
        binDif_top[y:y + h, x:x + w],
        binDif_bot[y:y + h, x:x + w],
    )


def _process_file(
    idx: int,
    path: Path,
    dm: np.ndarray,
    bm: np.ndarray,
    params: RegSegParams,
    top_dir: Path,
    topbw_dir: Path,
    bot_dir: Path,
    botbw_dir: Path,
    binDif_top: np.ndarray,
    binDif_bot: np.ndarray,
    roi: tuple[int, int, int, int] | None = None,
):
    """Process a single image file.

    Registration produces a binary overlap mask that constrains all
    subsequent processing.  Registered images and difference masks are
    multiplied by this mask, cropped to its bounding box and then passed to
    segmentation, ensuring that only valid pixels contribute to analysis.

    Returns the index, filename, current image, and processing results.
    """
    cur_full = imread_gray(path)
    dm_roi, bm_roi, binDif_top_roi, binDif_bot_roi = _apply_roi(
        dm, bm, binDif_top, binDif_bot, roi
    )
    # Register the full current image against the full DM but restrict the
    # optimisation to the ROI.  ``register_ecc`` returns arrays already cropped
    # to this ROI, which we then use for subtraction and segmentation.
    reg, mask = register_ecc(cur_full, dm, params, roi=roi)
    reg = reg * mask
    binDif_top_masked = binDif_top_roi * mask
    binDif_bot_masked = binDif_bot_roi * mask
    ys, xs = np.nonzero(mask)
    if ys.size == 0 or xs.size == 0:
        empty = np.zeros_like(mask)
        return idx, path.name, cur_full, empty, empty, [], []
    y1, y2 = ys.min(), ys.max() + 1
    x1, x2 = xs.min(), xs.max() + 1
    reg = reg[y1:y2, x1:x2]
    bm_crop = bm_roi[y1:y2, x1:x2]
    binDif_top_crop = binDif_top_masked[y1:y2, x1:x2]
    binDif_bot_crop = binDif_bot_masked[y1:y2, x1:x2]
    mask_crop = mask[y1:y2, x1:x2]
    binReg_top = cv2.subtract(reg, bm_crop)
    topDiff = cv2.subtract(clahe_equalize(binDif_top_crop), clahe_equalize(binReg_top))
    topDiff = topDiff * mask_crop
    cv2.imwrite(str(top_dir / path.name), to_uint8(topDiff))
    topBW = segment_image(topDiff, params, mask=mask_crop)
    cv2.imwrite(str(topbw_dir / path.name), to_uint8(topBW))
    areas_top = connected_component_areas(topBW)
    areas_top.sort(reverse=True)
    comp_bm_roi = complement(bm_roi)
    comp_bm_crop = comp_bm_roi[y1:y2, x1:x2]
    binReg_bot = cv2.subtract(reg, comp_bm_crop)
    botDiff = cv2.subtract(clahe_equalize(binDif_bot_crop), clahe_equalize(binReg_bot))
    botDiff = botDiff * mask_crop
    cv2.imwrite(str(bot_dir / path.name), to_uint8(botDiff))
    botBW = segment_image(botDiff, params, mask=mask_crop)
    cv2.imwrite(str(botbw_dir / path.name), to_uint8(botBW))
    areas_bot = connected_component_areas(botBW)
    areas_bot.sort(reverse=True)
    return idx, path.name, cur_full, topDiff, botDiff, areas_top, areas_bot


class ProcessorWorker(QObject):
    """Worker object that processes images in a separate thread."""

    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    imagePreviews = pyqtSignal(np.ndarray, np.ndarray, np.ndarray)
    topPreview = pyqtSignal(np.ndarray)
    bottomPreview = pyqtSignal(np.ndarray)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(
        self,
        in_dir: Path | str,
        out_dir: Path | str,
        dm_path: Path | str | None,
        bm_path: Path | str,
        params: RegSegParams,
        files: Sequence[Path],
        dm_roi: tuple[int, int, int, int] | None = None,
    ):
        super().__init__()
        self.in_dir = Path(in_dir)
        self.out_dir = Path(out_dir)
        self.dm_path = Path(dm_path) if dm_path is not None else None
        self.bm_path = Path(bm_path)
        self.params = params
        self.files = [Path(f) for f in files]
        self.dm_roi = dm_roi
        self._stop = False
        self._pause = False
        self._mutex = QMutex()
        self._wait_condition = QWaitCondition()

    def stop(self) -> None:
        """Request the worker thread to stop processing."""
        self._mutex.lock()
        try:
            self._stop = True
            self._wait_condition.wakeAll()
        finally:
            self._mutex.unlock()

    def toggle_pause(self, value: bool) -> None:
        """Pause or resume processing."""
        self._mutex.lock()
        try:
            self._pause = value
            if not self._pause:
                self._wait_condition.wakeAll()
        finally:
            self._mutex.unlock()

    def _check_pause_stop(self) -> None:
        """Wait while paused and raise if a stop was requested."""
        self._mutex.lock()
        try:
            while self._pause and not self._stop:
                self._wait_condition.wait(self._mutex)
            if self._stop:
                raise RuntimeError("Processing stopped by user.")
        finally:
            self._mutex.unlock()

    @pyqtSlot()
    def run(self) -> None:
        """Entry point for the worker thread."""
        top_wb = bot_wb = None
        try:
            files = self.files
            if not files:
                msg = f"No image files found in {self.in_dir}"
                logger.warning(msg)
                self.error.emit(msg)
                return
            sample = imread_gray(files[0])
            full_shape = sample.shape
            if self.dm_path is not None and self.dm_path.is_file():
                dm_raw = imread_gray(self.dm_path)
            else:
                # If no difference mask was provided, use a blank image so that
                # processing can still continue.  This mirrors the behaviour of
                # a zero DM on disk.
                dm_raw = np.zeros(full_shape, dtype=sample.dtype)
            bm_raw = imread_gray(self.bm_path)
            if self.dm_roi is not None:
                x, y, w, h = self.dm_roi
                dm = np.zeros(full_shape, dtype=dm_raw.dtype)
                bm = np.zeros(full_shape, dtype=bm_raw.dtype)
                dm[y:y + h, x:x + w] = dm_raw
                bm[y:y + h, x:x + w] = bm_raw
                roi = (x, y, w, h)
            else:
                dm = dm_raw
                bm = bm_raw
                roi = None
            self.imagePreviews.emit(dm, bm, dm)
            top_dir = self.out_dir / "top"
            topbw_dir = self.out_dir / "topBW"
            bot_dir = self.out_dir / "bottom"
            botbw_dir = self.out_dir / "bottomBW"
            for d in (top_dir, topbw_dir, bot_dir, botbw_dir):
                ensure_dir(d)
            binDif_top = cv2.subtract(dm, bm)
            binDif_bot = cv2.subtract(dm, complement(bm))
            total = len(files)
            top_xlsx = self.out_dir / "top.xlsx"
            bot_xlsx = self.out_dir / "bottom.xlsx"
            if top_xlsx.exists():
                top_wb = load_workbook(top_xlsx)
                top_ws = top_wb.active
            else:
                top_wb = Workbook()
                top_ws = top_wb.active
            if bot_xlsx.exists():
                bot_wb = load_workbook(bot_xlsx)
                bot_ws = bot_wb.active
            else:
                bot_wb = Workbook()
                bot_ws = bot_wb.active
            logger.info("Processing %d files in %s", total, self.in_dir)
            results_top: dict[int, list[int]] = {}
            results_bot: dict[int, list[int]] = {}
            with ThreadPoolExecutor() as executor:
                futures = [
                    executor.submit(
                        _process_file,
                        idx,
                        path,
                        dm,
                        bm,
                        self.params,
                        top_dir,
                        topbw_dir,
                        bot_dir,
                        botbw_dir,
                        binDif_top,
                        binDif_bot,
                        roi,
                    )
                    for idx, path in enumerate(files, start=1)
                ]
                completed = 0
                for future in as_completed(futures):
                    self._check_pause_stop()
                    (
                        idx,
                        fname,
                        cur,
                        topDiff,
                        botDiff,
                        areas_top,
                        areas_bot,
                    ) = future.result()
                    results_top[idx] = areas_top
                    results_bot[idx] = areas_bot
                    self.status.emit(f"Processing {idx}/{total}: {fname}")
                    self.imagePreviews.emit(dm, bm, cur)
                    # Emit previews only after both diffs are ready
                    self.topPreview.emit(topDiff)
                    self.bottomPreview.emit(botDiff)
                    completed += 1
                    progress_pct = int(completed / total * 100)
                    self.progress.emit(progress_pct)
            for idx in range(1, total + 1):
                areas_top = results_top.get(idx, [])
                for row, area in enumerate(areas_top, start=1):
                    top_ws.cell(row=row, column=idx, value=int(area))
                areas_bot = results_bot.get(idx, [])
                for row, area in enumerate(areas_bot, start=1):
                    bot_ws.cell(row=row, column=idx, value=int(area))
            self.status.emit("Done.")
            self.progress.emit(100)
        except Exception as exc:
            logger.exception("Processing failed")
            self.error.emit(str(exc))
        finally:
            if top_wb is not None:
                top_wb.save(top_xlsx)
            if bot_wb is not None:
                bot_wb.save(bot_xlsx)
            self.finished.emit()
